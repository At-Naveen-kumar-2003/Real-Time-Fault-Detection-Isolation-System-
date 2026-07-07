# =============================================================================
#  generate_dataset_excel.py
#  Run in PyCharm: right-click → Run 'generate_dataset_excel'
#
#  Outputs:
#    saved_models/training_data.csv          (for ML training)
#    saved_models/training_data.xlsx         (for GitHub / Excel viewing)
#
#  Excel sheets:
#    1. Dataset      — all samples (current, voltage, temp, gas, label …)
#    2. Summary      — class counts, percentages, feature stats
#    3. Class Info   — label names, descriptions, fault type
#    4. Feature Info — all 17 feature descriptions
# =============================================================================

import os
import sys
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side, GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

# ── Config (inline so file runs standalone in PyCharm) ────────────────────────
MODEL_DIR = "saved_models"
os.makedirs(MODEL_DIR, exist_ok=True)

FEATURE_NAMES = [
    "current", "voltage", "temperature", "gas", "power",
    "delta_current", "delta_voltage", "delta_temp", "delta_gas",
    "roll_std_current", "roll_std_voltage", "roll_std_temp", "roll_std_gas",
    "physics_error", "z_current", "z_voltage", "cross_delta_temp",
]
N_CLASSES = 11
FAULT_LABELS = {
    0: "Normal",           1: "Overcurrent",      2: "Overvoltage",
    3: "Undervoltage",     4: "Overtemperature",  5: "Gas Leak",
    6: "Spoofing Attack",  7: "Replay Attack",    8: "Gradual Drift",
    9: "Pulse Attack",     10: "Physics Attack",
}
FAULT_TYPE = {
    0: "Normal",     1: "Genuine Fault", 2: "Genuine Fault",
    3: "Genuine Fault", 4: "Genuine Fault", 5: "Genuine Fault",
    6: "Cyber Attack", 7: "Cyber Attack",  8: "Cyber Attack",
    9: "Cyber Attack", 10: "Cyber Attack",
}
DESCRIPTIONS = {
    0:  "All sensors in normal operating range",
    1:  "Current exceeds safe limit — hardware fault",
    2:  "Voltage above maximum threshold — hardware fault",
    3:  "Voltage below minimum threshold — hardware fault",
    4:  "Temperature spike — hardware overheating fault",
    5:  "Gas ppm exceeds safe limit — gas leak detected",
    6:  "Node 3 injects abrupt fake sensor values",
    7:  "Node 3 replays frozen normal values repeatedly",
    8:  "Values slowly creep toward fault threshold",
    9:  "Brief spikes interspersed with normal readings",
    10: "Reported power does not match V × I (physics violated)",
}
FEATURE_DESCRIPTIONS = {
    "current":          "Raw current reading (A)",
    "voltage":          "Raw voltage reading (V)",
    "temperature":      "Raw temperature reading (°C)",
    "gas":              "Raw gas concentration (ppm)",
    "power":            "Raw power reading (W)",
    "delta_current":    "Rate of change of current (A/step)",
    "delta_voltage":    "Rate of change of voltage (V/step)",
    "delta_temp":       "Rate of change of temperature (°C/step)",
    "delta_gas":        "Rate of change of gas (ppm/step)",
    "roll_std_current": "Rolling std of current over 5 steps",
    "roll_std_voltage": "Rolling std of voltage over 5 steps",
    "roll_std_temp":    "Rolling std of temperature over 5 steps",
    "roll_std_gas":     "Rolling std of gas over 5 steps",
    "physics_error":    "|P - V×I| / (V×I + ε) — physics violation",
    "z_current":        "Z-score of current vs rolling baseline",
    "z_voltage":        "Z-score of voltage vs rolling baseline",
    "cross_delta_temp": "|temp - normal_temp| — deviation from baseline",
}

# ── Dataset generation ─────────────────────────────────────────────────────────
np.random.seed(42)
NORMAL   = dict(curr=2.0, volt=11.0, temp=32.0, gas=100.0)
PROFILES = {
    0:  dict(curr=2.0,  volt=11.0, temp=32.0, gas=100.0),
    1:  dict(curr=6.8,  volt=10.2, temp=38.0, gas=110.0),
    2:  dict(curr=2.2,  volt=14.8, temp=35.0, gas=105.0),
    3:  dict(curr=1.8,  volt=7.2,  temp=30.0, gas=95.0),
    4:  dict(curr=2.3,  volt=11.0, temp=82.0, gas=130.0),
    5:  dict(curr=2.1,  volt=11.0, temp=33.0, gas=680.0),
    6:  dict(curr=6.8,  volt=10.2, temp=38.0, gas=110.0),
    7:  dict(curr=2.0,  volt=11.0, temp=32.0, gas=100.0),
    8:  dict(curr=6.8,  volt=10.2, temp=38.0, gas=110.0),
    9:  dict(curr=7.0,  volt=14.5, temp=80.0, gas=650.0),
    10: dict(curr=2.5,  volt=11.0, temp=33.0, gas=105.0),
}


def _rolling_std(arr, w=5):
    return np.array([arr[max(0, i - w):i + 1].std() for i in range(len(arr))])


def _make_sequence(cls, n):
    p = PROFILES[cls]
    if cls == 7:
        curr = np.full(n, p["curr"]) + np.random.normal(0, 0.002, n)
        volt = np.full(n, p["volt"]) + np.random.normal(0, 0.002, n)
        temp = np.full(n, p["temp"]) + np.random.normal(0, 0.01,  n)
        gas  = np.full(n, p["gas"])  + np.random.normal(0, 0.1,   n)
    elif cls == 8:
        t    = np.linspace(0, 1, n)
        curr = NORMAL["curr"] + t*(p["curr"]-NORMAL["curr"]) + np.random.normal(0, 0.03, n)
        volt = NORMAL["volt"] + t*(p["volt"]-NORMAL["volt"]) + np.random.normal(0, 0.03, n)
        temp = NORMAL["temp"] + t*(p["temp"]-NORMAL["temp"]) + np.random.normal(0, 0.30, n)
        gas  = NORMAL["gas"]  + t*(p["gas"] -NORMAL["gas"])  + np.random.normal(0, 2.00, n)
    elif cls == 9:
        curr = np.random.normal(NORMAL["curr"], 0.20, n)
        volt = np.random.normal(NORMAL["volt"], 0.30, n)
        temp = np.random.normal(NORMAL["temp"], 2.00, n)
        gas  = np.random.normal(NORMAL["gas"],  15.0, n)
        idx  = np.arange(0, n, 10)
        curr[idx] = p["curr"] + np.random.normal(0, 0.2, len(idx))
        volt[idx] = p["volt"] + np.random.normal(0, 0.2, len(idx))
        temp[idx] = p["temp"] + np.random.normal(0, 1.0, len(idx))
        gas[idx]  = p["gas"]  + np.random.normal(0, 10,  len(idx))
    else:
        noise = 0.05
        curr = np.random.normal(p["curr"], noise*max(p["curr"],0.5), n).clip(0.1, 15.0)
        volt = np.random.normal(p["volt"], noise*max(p["volt"],1.0), n).clip(0.5, 20.0)
        temp = np.random.normal(p["temp"], noise*max(p["temp"],5.0), n).clip(5.0, 110.0)
        gas  = np.random.normal(p["gas"],  noise*max(p["gas"], 20.), n).clip(20,  1000)

    pwr = curr * volt
    if cls == 10:
        pwr = curr * volt * np.random.uniform(1.25, 1.65, n)

    d_curr = np.diff(curr, prepend=curr[0])
    d_volt = np.diff(volt, prepend=volt[0])
    d_temp = np.diff(temp, prepend=temp[0])
    d_gas  = np.diff(gas,  prepend=gas[0])
    if cls == 6:
        d_curr[0] = p["curr"] - NORMAL["curr"]
        d_volt[0] = p["volt"] - NORMAL["volt"]
        d_temp[0] = p["temp"] - NORMAL["temp"]
        d_gas[0]  = p["gas"]  - NORMAL["gas"]

    rs_curr = _rolling_std(curr)
    rs_volt = _rolling_std(volt)
    rs_temp = _rolling_std(temp)
    rs_gas  = _rolling_std(gas)
    phys_err = np.abs(pwr - curr*volt) / (curr*volt + 1e-6)
    b    = max(5, n // 5)
    z_c  = (curr - curr[:b].mean()) / max(curr[:b].std(), 0.01)
    z_v  = (volt - volt[:b].mean()) / max(volt[:b].std(), 0.01)
    cross = np.abs(temp - NORMAL["temp"])

    return np.column_stack([
        curr, volt, temp, gas, pwr,
        d_curr, d_volt, d_temp, d_gas,
        rs_curr, rs_volt, rs_temp, rs_gas,
        phys_err, z_c, z_v, cross,
    ])


def generate_df(n_per_class=1500):
    all_X, all_y = [], []
    for cls in range(N_CLASSES):
        n = n_per_class * 5 if cls == 0 else n_per_class
        all_X.append(_make_sequence(cls, n))
        all_y.append(np.full(n, cls, dtype=int))
    X = np.vstack(all_X)
    y = np.concatenate(all_y)
    idx = np.random.permutation(len(y))
    df = pd.DataFrame(X[idx], columns=FEATURE_NAMES)
    df["label"] = y[idx]
    df.insert(0, "sample_id", range(1, len(df)+1))
    df.insert(df.columns.get_loc("label")+1, "fault_name",
              df["label"].map(FAULT_LABELS))
    df.insert(df.columns.get_loc("fault_name")+1, "fault_type",
              df["label"].map(FAULT_TYPE))
    return df


# ── Excel styling helpers ──────────────────────────────────────────────────────
def hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

CLASS_COLORS = {
    0:"C6EFCE",  1:"FFC7CE",  2:"FFC7CE",  3:"FFEB9C",
    4:"FFC7CE",  5:"FFC7CE",  6:"E2EFDA",  7:"E2EFDA",
    8:"E2EFDA",  9:"E2EFDA",  10:"E2EFDA",
}
CLASS_TEXT_COLORS = {
    0:"275A26",  1:"9C0006",  2:"9C0006",  3:"7D6608",
    4:"9C0006",  5:"9C0006",  6:"375623",  7:"375623",
    8:"375623",  9:"375623",  10:"375623",
}


# ══════════════════════════════════════════════════════════════════════════════
def build_excel(df):
    wb = Workbook()

    # ── Sheet 1: Dataset ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Dataset"

    # Header row
    headers = list(df.columns)
    for c, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        cell.fill      = hdr_fill("1F4E79")
        cell.alignment = center()
        cell.border    = thin_border()

    # Data rows — write in chunks for speed
    print("  Writing dataset sheet ...")
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        label = row[df.columns.get_loc("label") + 0]   # +0 because no index
        # Actually get the label from the label column position
        label_col = list(df.columns).index("label")
        lbl = int(row[label_col])
        fill_hex   = CLASS_COLORS.get(lbl, "FFFFFF")
        text_hex   = CLASS_TEXT_COLORS.get(lbl, "000000")

        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            cell.font      = Font(size=9, name="Arial", color=text_hex)
            cell.border    = thin_border()
            cell.alignment = left()
            # Color the fault_name and fault_type columns
            col_name = headers[c_idx - 1]
            if col_name in ("fault_name", "fault_type", "label"):
                cell.fill = hdr_fill(fill_hex)
            # Round floats
            if isinstance(value, float):
                cell.number_format = "0.0000"

    # Freeze top row, auto-filter
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = ws1.dimensions

    # Column widths
    col_widths = {"sample_id": 10, "label": 8, "fault_name": 18,
                  "fault_type": 14}
    for c, h in enumerate(headers, 1):
        ws1.column_dimensions[get_column_letter(c)].width = col_widths.get(h, 14)

    # Row height
    ws1.row_dimensions[1].height = 28

    # Conditional formatting — color scale on "current" column
    curr_col = get_column_letter(headers.index("current") + 1)
    last_row = len(df) + 1
    ws1.conditional_formatting.add(
        f"{curr_col}2:{curr_col}{last_row}",
        ColorScaleRule(
            start_type="min", start_color="63BE7B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="F8696B"
        )
    )

    # ── Sheet 2: Summary ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False

    title_font = Font(bold=True, size=14, name="Arial", color="1F4E79")
    ws2["A1"] = "Edge AI Fault Detection — Dataset Summary"
    ws2["A1"].font = title_font
    ws2.merge_cells("A1:G1")
    ws2["A1"].alignment = center()
    ws2.row_dimensions[1].height = 30

    ws2["A2"] = f"Total samples: {len(df):,}   |   Classes: {N_CLASSES}   |   Features: {len(FEATURE_NAMES)}"
    ws2["A2"].font = Font(size=10, name="Arial", color="595959", italic=True)
    ws2.merge_cells("A2:G2")
    ws2["A2"].alignment = center()

    # Class distribution table
    sum_headers = ["Class ID", "Class Name", "Fault Type", "Samples", "% of Total",
                   "Min Current", "Max Current"]
    for c, h in enumerate(sum_headers, 1):
        cell = ws2.cell(row=4, column=c, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        cell.fill      = hdr_fill("2E75B6")
        cell.alignment = center()
        cell.border    = thin_border()

    class_counts = df["label"].value_counts().sort_index()
    total = len(df)
    for r, cls in enumerate(range(N_CLASSES), 5):
        cnt = class_counts.get(cls, 0)
        pct = cnt / total * 100
        sub = df[df["label"] == cls]
        row_data = [
            cls,
            FAULT_LABELS[cls],
            FAULT_TYPE[cls],
            cnt,
            f"{pct:.1f}%",
            round(sub["current"].min(), 3) if len(sub) else 0,
            round(sub["current"].max(), 3) if len(sub) else 0,
        ]
        fill_hex = CLASS_COLORS.get(cls, "FFFFFF")
        text_hex = CLASS_TEXT_COLORS.get(cls, "000000")
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.font      = Font(size=9, name="Arial", color=text_hex)
            cell.fill      = hdr_fill(fill_hex)
            cell.alignment = center()
            cell.border    = thin_border()

    # Totals row
    tot_row = 5 + N_CLASSES
    ws2.cell(row=tot_row, column=1, value="TOTAL").font = Font(bold=True, name="Arial", size=9)
    ws2.cell(row=tot_row, column=4, value=f"=SUM(D5:D{tot_row-1})")
    ws2.cell(row=tot_row, column=5, value="100.0%")
    for c in range(1, 8):
        cell = ws2.cell(row=tot_row, column=c)
        cell.fill   = hdr_fill("D6E4F0")
        cell.font   = Font(bold=True, name="Arial", size=9)
        cell.border = thin_border()
        cell.alignment = center()

    ws2.row_dimensions[4].height = 22
    col_ws2 = [10, 20, 14, 10, 12, 14, 14]
    for i, w in enumerate(col_ws2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Feature statistics table
    stat_row = tot_row + 3
    ws2.cell(row=stat_row, column=1, value="Feature Statistics (full dataset)").font = \
        Font(bold=True, size=11, name="Arial", color="1F4E79")
    ws2.merge_cells(f"A{stat_row}:G{stat_row}")

    stat_headers = ["Feature", "Min", "Max", "Mean", "Std Dev", "Type", "Unit"]
    for c, h in enumerate(stat_headers, 1):
        cell = ws2.cell(row=stat_row+1, column=c, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=9, name="Arial")
        cell.fill      = hdr_fill("375623")
        cell.alignment = center()
        cell.border    = thin_border()

    feature_units = {
        "current":"A","voltage":"V","temperature":"°C","gas":"ppm","power":"W",
        "delta_current":"A/step","delta_voltage":"V/step",
        "delta_temp":"°C/step","delta_gas":"ppm/step",
        "roll_std_current":"A","roll_std_voltage":"V",
        "roll_std_temp":"°C","roll_std_gas":"ppm",
        "physics_error":"ratio","z_current":"σ","z_voltage":"σ","cross_delta_temp":"°C",
    }
    feature_types = {
        "current":"raw","voltage":"raw","temperature":"raw","gas":"raw","power":"raw",
        "delta_current":"engineered","delta_voltage":"engineered",
        "delta_temp":"engineered","delta_gas":"engineered",
        "roll_std_current":"engineered","roll_std_voltage":"engineered",
        "roll_std_temp":"engineered","roll_std_gas":"engineered",
        "physics_error":"engineered","z_current":"engineered",
        "z_voltage":"engineered","cross_delta_temp":"engineered",
    }
    alt_colors = ["F2F2F2", "FFFFFF"]
    for i, feat in enumerate(FEATURE_NAMES):
        r = stat_row + 2 + i
        vals = [
            feat,
            round(float(df[feat].min()), 4),
            round(float(df[feat].max()), 4),
            round(float(df[feat].mean()), 4),
            round(float(df[feat].std()), 4),
            feature_types.get(feat, "raw"),
            feature_units.get(feat, ""),
        ]
        fill_hex = alt_colors[i % 2]
        for c, val in enumerate(vals, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.font      = Font(size=9, name="Arial")
            cell.fill      = hdr_fill(fill_hex)
            cell.alignment = center()
            cell.border    = thin_border()
            if isinstance(val, float):
                cell.number_format = "0.0000"

    # ── Sheet 3: Class Info ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("Class Info")
    ws3.sheet_view.showGridLines = False
    ws3["A1"] = "11-Class Fault / Attack Reference"
    ws3["A1"].font = Font(bold=True, size=13, name="Arial", color="1F4E79")
    ws3.merge_cells("A1:E1")
    ws3["A1"].alignment = center()
    ws3.row_dimensions[1].height = 28

    ci_headers = ["Class ID", "Label", "Fault Type", "Description", "Key Feature Signal"]
    key_signals = {
        0: "All features near baseline",
        1: "High current, delta_current spike",
        2: "High voltage, delta_voltage spike",
        3: "Low voltage",
        4: "High temp, delta_temp large",
        5: "High gas, delta_gas spike",
        6: "Huge delta on all channels (1st step)",
        7: "Near-zero roll_std (frozen values)",
        8: "Slowly rising delta features",
        9: "Periodic spikes in roll_std",
        10: "physics_error > 0.25 consistently",
    }
    for c, h in enumerate(ci_headers, 1):
        cell = ws3.cell(row=3, column=c, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        cell.fill      = hdr_fill("7030A0")
        cell.alignment = center()
        cell.border    = thin_border()
        ws3.row_dimensions[3].height = 22

    for r, cls in enumerate(range(N_CLASSES), 4):
        fill_hex = CLASS_COLORS.get(cls, "FFFFFF")
        text_hex = CLASS_TEXT_COLORS.get(cls, "000000")
        row_vals = [cls, FAULT_LABELS[cls], FAULT_TYPE[cls],
                    DESCRIPTIONS[cls], key_signals[cls]]
        for c, val in enumerate(row_vals, 1):
            cell = ws3.cell(row=r, column=c, value=val)
            cell.font      = Font(size=9, name="Arial", color=text_hex)
            cell.fill      = hdr_fill(fill_hex)
            cell.alignment = left() if c >= 4 else center()
            cell.border    = thin_border()
        ws3.row_dimensions[r].height = 20

    for col, w in zip("ABCDE", [10, 18, 14, 42, 36]):
        ws3.column_dimensions[col].width = w

    # ── Sheet 4: Feature Info ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("Feature Info")
    ws4.sheet_view.showGridLines = False
    ws4["A1"] = "17 Engineered Features — Reference"
    ws4["A1"].font = Font(bold=True, size=13, name="Arial", color="1F4E79")
    ws4.merge_cells("A1:D1")
    ws4["A1"].alignment = center()
    ws4.row_dimensions[1].height = 28

    fi_headers = ["#", "Feature Name", "Description", "Type"]
    for c, h in enumerate(fi_headers, 1):
        cell = ws4.cell(row=3, column=c, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        cell.fill      = hdr_fill("0070C0")
        cell.alignment = center()
        cell.border    = thin_border()
        ws4.row_dimensions[3].height = 22

    raw_feats = {"current","voltage","temperature","gas","power"}
    alt_colors = ["EBF3FB", "FFFFFF"]
    for i, feat in enumerate(FEATURE_NAMES, 1):
        r = i + 3
        ftype = "Raw sensor" if feat in raw_feats else "Engineered"
        row_vals = [i, feat, FEATURE_DESCRIPTIONS[feat], ftype]
        fill_hex = alt_colors[i % 2]
        for c, val in enumerate(row_vals, 1):
            cell = ws4.cell(row=r, column=c, value=val)
            cell.font      = Font(size=9, name="Arial")
            cell.fill      = hdr_fill(fill_hex)
            cell.alignment = center() if c in (1, 4) else left()
            cell.border    = thin_border()
        ws4.row_dimensions[r].height = 18

    for col, w in zip("ABCD", [6, 22, 52, 14]):
        ws4.column_dimensions[col].width = w

    return wb


# ══════════════════════════════════════════════════════════════════════════════
def main():
    print("=" * 60)
    print("  Edge AI — Dataset Generator + Excel Exporter")
    print("=" * 60)

    # 1. Generate
    print("\n[1/3] Generating 11-class dataset (1500 samples per class) ...")
    df = generate_df(n_per_class=1500)
    print(f"      Total samples: {len(df):,}")
    print(f"      Features     : {len(FEATURE_NAMES)}")
    print(f"      Classes      : {N_CLASSES}")

    # 2. Save CSV
    csv_path = os.path.join(MODEL_DIR, "training_data.csv")
    df.to_csv(csv_path, index=False)
    print(f"\n[2/3] CSV saved  → {csv_path}")

    # 3. Save Excel
    print("\n[3/3] Building Excel workbook ...")
    wb = build_excel(df)
    xlsx_path = os.path.join(MODEL_DIR, "training_data.xlsx")
    wb.save(xlsx_path)
    print(f"      Excel saved  → {xlsx_path}")

    print("\n" + "=" * 60)
    print("  Class breakdown:")
    for cls in range(N_CLASSES):
        cnt = (df["label"] == cls).sum()
        print(f"    {cls:>2}  {FAULT_LABELS[cls]:<22}  {cnt:>6,} samples")

    print("\n  Sheets in Excel:")
    print("    1. Dataset    — all samples with color-coded labels")
    print("    2. Summary    — class counts + feature statistics")
    print("    3. Class Info — fault descriptions + key signals")
    print("    4. Feature Info — 17 feature reference table")
    print("=" * 60)
    print(f"\n  Done! Open: {xlsx_path}")


if __name__ == "__main__":
    main()